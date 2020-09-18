from docx import Document
from win32com import client
import datetime
from openpyxl import load_workbook
import os

'''
原文链接：https://mp.weixin.qq.com/s?__biz=MzI1MTUyMjc1Mg==&mid=2247485450&idx=1&sn=37461f609a0e89c9e205abf60683b224&chksm=e9f0f3a5de877ab3c52d282aac8382a85a2e2558a1d5ecdbdfbf6731a7ea2fea426c49d0aa56&scene=21#wechat_redirect
'''
# TODO: 首先我们使用Python对该Excel进行解析
# 获取桌面的路径
def GetDesktopPath():
    return os.path.join(os.path.expanduser("~"), 'Desktop')
path = GetDesktopPath() + '/资料/'  # 形成文件夹的路径便后续重复使用 
workbook = load_workbook(filename=path + '数据.xlsx')
sheet = workbook.active  # 获取当前页
# 可以用代码获取数据范围，如果要批处理循环迭代也方便
# 获取有数据范围
print(sheet.dimensions)
# A1:W10

# TODO: 利用openpyxl读取单元格有以下几种用法∏
cells = sheet['A1:A4']  # 返回A1-A4的4个单元格
cells = sheet['A']  # 获取A列
cells = sheet['A:C']  # 获取A-C列
cells = sheet[5]  # 获取第5行
# 注意如果是上述用cells获取返回的是嵌套元祖
for cell in cells:
    print(cell[0].value)  # 遍历cells依然需要取出元祖中元素才可以获取值
# 获取一个范围的所有cell
# 也可以用iter_col返回列
for row in sheet.iter_rows(min_row=1, max_row=3, min_col=2, max_col=4):
    for cell in row:
        print(cell.value)

#明白了原理我们就可以解析获取Excel中的数据了
# SQE
SQE = sheet['Q2'].value
# 供应商&制造商
supplier = sheet['G2'].value
# 采购单号
C2_10 = sheet['C2:C10']  # 返回cell.tuple对象
# 利用列表推导式后面同理
vC2_10 = [str(cell[0].value) for cell in C2_10]
# 用set简易去重后用,连接，填word表用
order_num = ','.join(set(vC2_10))
# 用set简易去重后用&连接，word文件名命名使用
order_num_title = '&'.join(set(vC2_10))
# 产品型号
T2_10 = sheet['T2:T10']
vT2_10 = [str(cell[0].value) for cell in T2_10]
ptype = ','.join(set(vT2_10))
# 产品描述
P2_10 = sheet['P2:P10']
vP2_10 = [str(cell[0].value) for cell in P2_10]
info = ','.join(set(vP2_10))
info_title = '&'.join(set(vP2_10))
# 日期
# 用datetime库获取今日时间以及相应格式化
today = datetime.datetime.today()
time = today.strftime('%Y年%m月%d日')
# 验货数量
V2_10 = sheet['V2:V10']
vV2_10 = [int(cell[0].value) for cell in V2_10]
total_num = sum(vV2_10)  # 计算总数量
# 验货箱数
W2_10 = sheet['W2:W10']
vW2_10 = [int(cell[0].value) for cell in W2_10]
box_num = sum(vW2_10)
# 生成最终需要的word文件名
title = f'{order_num_title}-{supplier}-{total_num}-{info_title}-{time}-验货报告'
print(title)


'''
通过上面的代码，我们就成功的从Excel中提取出来数据，这样Excel部分就结束了，接下来进行word的填表啦，由于这里我们默认读取的word是.docx格式的，实际上读者的需求是.doc格式文件，所以windows用户可以用如下代码批量转化doc，前提是安装好win32com
'''
# pip install pypiwin32
docx_path = path + '模板.docx'
# doc转docx的函数


def doc2docx(doc_path, docx_path):
    word = client.Dispatch("Word.Application")
    doc = word.Documents.Open(doc_path)
    doc.SaveAs(docx_path, 16)
    doc.Close()
    word.Quit()
    print('\n doc文件已转换为docx \n')


if not os.path.exists(docx_path):
    doc2docx(docx_path[:-1], docx_path)

#不过在Mac下暂时没有好的解决策略，如果有思路欢迎交流，好了有docx格式文件后我们继续操作Word部分
docx_path = path + '模板.docx'
# 实例化
document = Document(docx_path)
# 读取word中的所有表格
tables = document.tables
# print(len(tables))
# 15


# 确定好每个表格数后即可进行相应的填报操作，table的用法和openpyxl中非常类似，注意索引和原生python一样都是从0开始
tables[0].cell(1, 1).text = SQE
tables[1].cell(1, 1).text = supplier
tables[1].cell(2, 1).text = supplier
tables[1].cell(3, 1).text = ptype
tables[1].cell(4, 1).text = info
tables[1].cell(5, 1).text = order_num
tables[1].cell(7, 1).text = time


# 我们继续用Python填写下一个表格
for i in range(2, 11):
    tables[6].cell(i, 0).text = str(sheet[f'T{i}'].value)
    tables[6].cell(i, 1).text = str(sheet[f'P{i}'].value)
    tables[6].cell(i, 2).text = str(sheet[f'C{i}'].value)
    tables[6].cell(i, 4).text = str(sheet[f'V{i}'].value)
    tables[6].cell(i, 5).text = str(sheet[f'V{i}'].value)
    tables[6].cell(i, 6).text = '0'
    tables[6].cell(i, 7).text = str(sheet[f'W{i}'].value)
    tables[6].cell(i, 8).text = '0'

tables[6].cell(12, 4).text = str(total_num)
tables[6].cell(12, 5).text = str(total_num)
tables[6].cell(12, 7).text = str(box_num)

# 写出文件
document.save(path + f'{title}.docx')
print('\n文件已生成')
