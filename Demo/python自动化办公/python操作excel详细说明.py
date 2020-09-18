import openpyxl

#1. 获取工作对象
workbook = openpyxl.load_workbook("C:\\Users\\GG257\\Desktop\\a.xlsx")
print(workbook)


#2. 获取工作表明
# shenames = workbook.get_sheet_names()
# print(shenames)  # ['Sheet1', 'Sheet2']
#如果上面的已经被弃用，可以使用下面的方法
shenames2 = workbook.sheetnames
print(shenames2)  # ['Sheet1', 'Sheet2']


#3. 获取工作表对象
#获得工作簿的表名后，就可以获得表对象
# worksheet = workbook.get_sheet_by_name('Sheet1')
# print(worksheet)  # <Worksheet "Sheet1">
#也可提使用下面的比较新的句子
# worksheet2 = workbook['Sheet1']
# print(worksheet2)
#还可以通过下面的来实现
worksheet = workbook[shenames2[0]]
print(worksheet)


#4. 根据索引方式获取工作表对象
# 选择活跃的表，默认是第一个
ws = workbook.active


#5. 按照行和列读取内容
# sheet.rows，这是一个生成器，里面是每一行数据，每一行数据由一个元组类型包裹
# sheet.columns，同上，里面是每一列数据。
# for row in worksheet.rows:
#     print(row)
#     for cell in row:
#         print(cell.value, end=" ")
#     print()

# for col in worksheet.columns:
#     for cell in col:
#         print(cell.value, end=" ")
#     print()


#6. 获取特定行或者列的数据
#我们可以想到的是用“索引”的方式，但是sheet.rows是生成器类型，不能使用索引。所以我们将其转换为list之后再使用索引，例如用list(sheet.rows)[3]来获取第四行的tuple对象。
#读取第四行
for cell in list(worksheet.rows)[3]:
    print(cell.value, end=' ')
print()
#读取第3列
for cell in list(worksheet.columns)[2]:
    print(cell.value, end=' ')
print()
#读取某一块的数据, 记住，worksheet默认是迭代器，去单独某一块需要使用到list，所以需要转译
for row in list(worksheet.rows)[0:2]:
    for cell in row[0:2]:
        print(cell.value, end=' ')
    print()
#或者使用数值标记读取
for i in range(1, 3):
    for j in range(1, 3):
        print(worksheet.cell(row=i, column=j).value, end=' ')
    print()


#7. 获得某一单元格的数据
content_A1 = worksheet['A1'].value
print(content_A1)
content_A1_1 = worksheet.cell(row=1, column=2).value
print(content_A1_1)


################################################################  写操作  ###
#创建工作簿和获取工作表
# 创建一个Workbook对象，相当于创建了一个Excel文件
workbook = openpyxl.Workbook()
##wb=openpyxl.Workbook(encoding='UTF-8')

#获取当前活跃的worksheet,默认就是第一个worksheet
worksheet = workbook.active
worksheet.title = 'mysheet'

#创建新的工作表
worksheet2 = workbook.create_sheet()  # 默认插入在工作簿末尾
# worksheet2 = workbook.create_sheet(0)  #默认插入在工作簿索引为0的位置上
worksheet2.title = 'New Title'

#将数据写入到工作表
Province = ['北京市', '天津市', '河北省', '山西省', '内蒙古自治区', '辽宁省',
            '吉林省', '黑龙江省', '上海市', '江苏省', '浙江省', '安徽省', '福建省',
            '江西省', '山东省', '河南省', '湖北省', '湖南省', '广东省', '广西壮族自治区',
            '海南省', '重庆市', '四川省', '贵州省', '云南省', '西藏自治区', '陕西省', '甘肃省',
            '青海省', '宁夏回族自治区', '新疆维吾尔自治区']
Income = ['5047.4', '3247.9', '1514.7', '1374.3', '590.7', '1499.5', '605.1', '654.9',
          '6686.0', '3104.8', '3575.1', '1184.1', '1855.5', '1441.3', '1671.5', '1022.7',
          '1199.2', '1449.6', '2906.2', '972.3', '555.7', '1309.9', '1219.5', '715.5', '441.8',
          '568.4', '848.3', '637.4', '653.3', '823.1', '254.1']
Project = ['各省市', '工资性收入', '家庭经营纯收入', '财产性收入', '转移性收入', '食品', '衣着',
           '居住', '家庭设备及服务', '交通和通讯', '文教、娱乐用品及服务', '医疗保健', '其他商品及服务']

#写入第一行数据,行和列都是从1开始计数
for i in range(len(Project)):
    worksheet.cell(1, i+1, Project[i])
#写入第一列数据，因为第一行已经有了，所以从第二列开始计数
for i in range(len(Province)):
    worksheet.cell(i+2, 1, Province[i])
#写入第二列数据
for i in range(len(Income)):
    worksheet.cell(i+2, 2, Income[i])


# 保存工作簿
worksheet.save(filename='path')


#修改已经存在的工作簿
workbook = openpyxl.load_workbook("path")
worksheet = workbook.worksheets[0]
#在第一列之前插入一列
worksheet.insert_cols(1)
for index, row in enumerate(worksheet.rows):
    if index == 0:
        row[0].value = "编号"
    else:
        row[0].value = index
worksheet.save(filename='path')

#修改特定单元格
worksheet.cell(2, 3, '0')
worksheet['B2'] = "北京"

#批量修改数据
taiwan = [32, "台湾省"]
worksheet.append(taiwan)
