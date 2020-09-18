# 使用openpyxl批量提取数据

from openpyxl import load_workbook, Workbook

#数据所在目录
path = '***'
#打开数据
workbook = load_workbook(path+'/电商婴儿数据.xlsx')
#打开工作表
sheet = workbook.active
buy_mount = sheet['F']
row_lst = []
for cell in buy_mount:
    if isinstance(cell.value, int) and cell.value>50:
        print(cell.row)
        row_lst.append(cell.row)

new_workbook = Workbook()
new_sheet = new_workbook.active

header = sheet[1]
header_lst = []
for cell in header:
    header_lst.append(cell.calue)
new_sheet.append(header_lst)

for row in row_lst:
    data_lst = []
    for cell in sheet[row]:
        data_lst.append(cell.value)
    new_sheet.append(data_lst)

new_wordbook.save(path+'/符合筛选条件的新表.xlsx')

import glob
path='*****'
for file in glob.glob(path+'/*'):
    pass

for file in glob.glob(path+'/*.xlsx'):
    pass

#完整代码
from openpyxl import load_workbook, Workbook
import glob
path = "***"
new_workbook = Workbook()
new_sheet = new_workbook.active

flag = 0
for file in glob.glob(path+'/*.xlsx'):
    workbook = load_workbook(file)
    sheet = workbook.active
    by_mount = sheet['F']
    row_lst = []
    for cell in buy_mount:
        if isinstance(cell.value, int) and cell.value > 50:
            print(cell.value)
            row_lst.append(cell.row)
    if not flag:
        header = sheet[1]
        header_lst = []
        for cell in header:
            header_lst.append(cell.value)
        new_sheet.append(header_lst)
        flag = 1

    for row in row_lst:
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)
new_workbok.save(path+'/符合筛选条件的新表.xlsx')
